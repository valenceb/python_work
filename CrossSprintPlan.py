from openpyxl import Workbook
import datetime
import csv
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("output/sample.xlsx")

if __name__ == '__main__':
	projectList = []
	sprintList = []
	with open('conf/project.csv') as f:
		empf = csv.reader(f)
		for emp in empf:
			projectList.append(emp)
	print (projectList)
	with open('conf/sprint.csv') as f:
		empf = csv.reader(f)
		for emp in empf:
			sprintList.append(emp)
	print (sprintList)