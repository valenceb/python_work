#!/usr/bin/env python
# -*- coding:utf-8 -*-
__author__ = "valenceb"

from openpyxl import Workbook
from excelStyle import getBodyStyle, getHeaderStyle, getProjHeaderStyle, getBlackFont
from crossSprintModal import Project, Sprint
import csv

if __name__ == '__main__':
    projectList = []
    sprintList = []
    wb = Workbook()
    wb.add_named_style(getHeaderStyle())
    wb.add_named_style(getProjHeaderStyle())
    wb.add_named_style(getBodyStyle())
    wb.add_named_style(getBlackFont())
    ws = wb.active
    ws['A1'] = "Project/Sprints"
    ws['A1'].style = 'header'
    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'B1'

    with open('conf/sprint.csv') as f:
        csvLines = csv.reader(f)
        i = 1
        for csvL in csvLines:
            sprintList.append(Sprint(''.join(csvL), i, ws))
            i += 1

    with open('conf/project.csv') as f:
        csvLines = csv.reader(f)
        i = 2
        for csvL in csvLines:
            proj = Project(''.join(csvL), sprintList, i, ws)
            if not proj.title:
                print('No such project.')
            else:
                projectList.append(proj)
            i += 1
    print(projectList)
    print(sprintList)
    wb.save("output/CrossSprintReport.xlsx")
    print('\nProcess Done.')
