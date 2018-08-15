from openpyxl import Workbook
#from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill,Color
from PIL import Image
wb = Workbook()
ws1 = wb.get_active_sheet()
im = Image.open("input/25.jpg")
print (im.size)
(image_width,image_height) = im.size
w=1
h=1
if image_width >= 500:
	image_width = int(image_width/2)
	w=2
if image_height >= 500:
	image_height = int(image_height/2)
	h=2
pix = im.load()
for row in range(2,image_height):    
    print ('row',row)
    for col in range(2,image_width):
        cell = ws1.cell(column=col, row=row, value='')
        pixColor = "FF%02X%02X%02X" % (pix[col*w-1-1,row*h-1-1][0],pix[col*w-1-1,row*h-1-1][1],pix[col*w-1-1,row*h-1-1][2])
        #print (pixColor)
        cell.fill=PatternFill(patternType='solid', fgColor=Color(rgb=pixColor))
        ws1.column_dimensions[cell.column].width = 1
    ws1.row_dimensions[row].height = 6
#for col in range(1,image_width):
#    ws1.column_dimensions[ws1.get_column_letter(col)].width = 1   
wb.save(filename = 'output/excelPicture.xlsx')
